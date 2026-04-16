from __future__ import annotations
import logging
import re
import shutil
from collections import defaultdict
from pathlib import Path
from typing import Callable
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet


def find_column(ws: Worksheet, name: str) -> int:
    """Return 1-based column index of the header named `name` in row 1."""
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value == name:
            return col
    raise ValueError(f"Column '{name}' not found in row 1")


def insert_calculated_column(
    ws: Worksheet,
    after_col_name: str,
    new_col_name: str,
    formula_fn: Callable[[float, int], float],
) -> None:
    """Insert a new column immediately after `after_col_name`.

    Writes `new_col_name` as header in row 1.
    For each data row (2..max_row), reads the source column value,
    applies formula_fn(val, row), fills result. ZeroDivisionError → 0.
    """
    source_idx = find_column(ws, after_col_name)
    insert_at = source_idx + 1
    ws.insert_cols(insert_at)
    ws.cell(1, insert_at).value = new_col_name
    # max_row was captured before insert (insert_cols doesn't change row count)
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, source_idx).value or 0
        try:
            result = formula_fn(float(val), row)  # type: ignore[arg-type]
        except (ZeroDivisionError, TypeError, ValueError):
            logging.warning("row %d: cannot compute %s from value %r, using 0", row, new_col_name, val)
            result = 0
        ws.cell(row, insert_at).value = result  # type: ignore[assignment]


def apply_calculated_columns(ws: Worksheet) -> None:
    """Insert 4 calculated columns into Sheet 4 in sequence.

    Re-scans column indices after each insertion to account for shifting.
    """
    # Column 1: 实际花费 = 花费 / 1.136
    insert_calculated_column(ws, "花费", "实际花费", lambda val, row: val / 1.136)

    # Column 2: 点击率 = 点击量 / 展现量
    # Capture 展现量 index now (after 实际花费 insertion shifted columns right of 花费)
    zx_col = find_column(ws, "展现量")
    insert_calculated_column(
        ws, "点击量", "点击率",
        lambda val, row, _zx=zx_col: val / (ws.cell(row, _zx).value or 1)
        if (ws.cell(row, _zx).value or 0) != 0 else 0
    )

    # Column 3: CPC = 实际花费 / 点击量
    # Re-scan after 点击率 insertion
    sjhf_col = find_column(ws, "实际花费")
    jl_col = find_column(ws, "点击量")
    insert_calculated_column(
        ws, "点击率", "CPC",
        lambda val, row, _jf=sjhf_col, _jl=jl_col: (
            (ws.cell(row, _jf).value or 0) / (ws.cell(row, _jl).value or 1)
            if (ws.cell(row, _jl).value or 0) != 0 else 0
        )
    )

    # Column 4: 实际成本 = 实际花费 / 留资人数
    # Re-scan after CPC insertion
    sjhf_col = find_column(ws, "实际花费")
    lzrs_col = find_column(ws, "留资人数")
    insert_calculated_column(
        ws, "留资成本", "实际成本",
        lambda val, row, _jf=sjhf_col, _lzrs=lzrs_col: (
            (ws.cell(row, _jf).value or 0) / (ws.cell(row, _lzrs).value or 1)
            if (ws.cell(row, _lzrs).value or 0) != 0 else 0
        )
    )


def sort_by_column(ws: Worksheet, col_name: str, descending: bool = True) -> None:
    """Sort data rows (row 2 to max_row) by the named column.

    Row 1 (header) is never moved. Entire rows move together.
    """
    sort_col = find_column(ws, col_name)
    max_col = ws.max_column
    max_row = ws.max_row

    # Read all data rows into memory
    data_rows = []
    for row in range(2, max_row + 1):
        row_values = [ws.cell(row, col).value for col in range(1, max_col + 1)]
        data_rows.append(row_values)

    # Sort by the target column; None values always go to the end
    none_rows = [r for r in data_rows if r[sort_col - 1] is None]
    value_rows = [r for r in data_rows if r[sort_col - 1] is not None]
    value_rows.sort(key=lambda r: r[sort_col - 1], reverse=descending)
    data_rows = value_rows + none_rows

    # Write sorted rows back
    for row_idx, row_values in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row_idx, col_idx).value = value


def group_and_merge(ws: Worksheet) -> None:
    """Insert 占比 column, group rows by 实际成本 ranges, merge and fill."""
    # Insert 占比 column after 互动成本
    hd_col = find_column(ws, "互动成本")
    zb_col = hd_col + 1
    ws.insert_cols(zb_col)
    ws.cell(1, zb_col).value = "占比"

    # Read 实际成本 values (column may have shifted after insert — re-scan)
    sjcb_col = find_column(ws, "实际成本")
    max_row = ws.max_row
    total_rows = max_row - 1  # row 1 is header

    if total_rows == 0:
        return

    # Partition into contiguous groups (data is pre-sorted descending)
    # Groups: high (≥90), medium (50-89), low (<50)
    groups = []
    current_range = None
    current_start = None

    for row in range(2, max_row + 1):
        val = ws.cell(row, sjcb_col).value or 0
        if val >= 90:
            rng = "high"
        elif val >= 50:
            rng = "medium"
        else:
            rng = "low"

        if rng != current_range:
            if current_range is not None:
                groups.append((current_range, current_start, row - 1))
            current_range = rng
            current_start = row

    if current_range is not None:
        groups.append((current_range, current_start, max_row))

    # Merge and fill each group
    for _rng, start_row, end_row in groups:
        count = end_row - start_row + 1
        pct = round(count / total_rows * 100)
        text = f"{count}/{pct}%"

        if start_row == end_row:
            # Single row — just write, no merge needed
            ws.cell(start_row, zb_col).value = text
        else:
            ws.merge_cells(
                start_row=start_row, start_column=zb_col,
                end_row=end_row, end_column=zb_col
            )
            ws.cell(start_row, zb_col).value = text


# Matches: =SUMIFS('SheetName'!<sum_col>:<sum_col>,'SheetName'!<src_key_col>:<src_key_col>,<tgt_key_col><row>)
# Groups: (1) sheet_name, (2) sum_col_letter, (3) src_key_col_letter, (4) tgt_key_col_letter
# The spec fixes the source sheet layout so src_key_col is always A, but we still
# capture it so an unexpected layout surfaces as a different sum rather than a
# silent aggregate from the wrong column.
SUMIFS_RE = re.compile(
    r"=SUMIFS\('([^']+)'!([A-Z]+):[A-Z]+,'[^']+'!([A-Z]+):[A-Z]+,([A-Z]+)\d+\)"
)

# Matches simple division formulas like =C2/F2, =C3/E3
# Groups: (1) numerator column letter, (2) denominator column letter
DIV_RE = re.compile(r"=([A-Z]+)\d+/([A-Z]+)\d+")


def resolve_formulas(wb: Workbook, ws: Worksheet) -> None:
    """Resolve SUMIFS and division formula strings in ws into numeric values.

    Three passes over row 2:
      1. SUMIFS — aggregate source sheet values by key column
      2. Division — compute =X/Y using values already written by pass 1
      3. Residual sweep — any remaining formula strings replaced with 0

    Passes 2 and 3 run even when no SUMIFS formulas are present, so division
    formulas and stray formula strings are still cleaned up on plain sheets.
    """
    # --- Phase 1: Scan row 2 to find SUMIFS formula columns ---
    # sumifs_cols maps: target_col_idx -> (sheet_name, sum_col_idx, src_key_col_idx, tgt_key_col_idx)
    sumifs_cols: dict[int, tuple[str, int, int, int]] = {}
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(2, col).value
        if not isinstance(cell_val, str):
            continue
        m = SUMIFS_RE.match(cell_val)
        if m:
            sheet_name = m.group(1)
            sum_col_idx = column_index_from_string(m.group(2))
            src_key_col_idx = column_index_from_string(m.group(3))
            tgt_key_col_idx = column_index_from_string(m.group(4))
            sumifs_cols[col] = (sheet_name, sum_col_idx, src_key_col_idx, tgt_key_col_idx)

    if sumifs_cols:
        # All SUMIFS formulas share the same source sheet and key layout; take the first
        source_meta = next(iter(sumifs_cols.values()))
        sheet_name = source_meta[0]
        src_key_col_idx = source_meta[2]
        tgt_key_col_idx = source_meta[3]

        # --- Phase 2: Load source sheet ---
        if sheet_name not in wb.sheetnames:
            logging.warning("resolve_formulas: source sheet '%s' not found; filling 0", sheet_name)
            for row in range(2, ws.max_row + 1):
                for col in sumifs_cols:
                    ws.cell(row, col).value = 0
        else:
            ws_src = wb[sheet_name]

            # --- Phase 3: Build aggregation dict: key -> {target_col -> sum} ---
            aggregated: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))

            for src_row in range(2, ws_src.max_row + 1):
                key_val = ws_src.cell(src_row, src_key_col_idx).value
                # Treat missing keys (None or empty string) as unmatchable — otherwise a
                # target row with key=None would silently collide with a source row whose
                # key is "".
                if key_val is None or key_val == "":
                    continue
                key_str = str(key_val)
                for tgt_col, (_sn, sum_col_idx, _skc, _tkc) in sumifs_cols.items():
                    raw = ws_src.cell(src_row, sum_col_idx).value
                    try:
                        aggregated[key_str][tgt_col] += float(raw or 0)
                    except (TypeError, ValueError):
                        pass  # non-numeric source value — skip

            # --- Phase 4: Fill target rows ---
            for row in range(2, ws.max_row + 1):
                key_val = ws.cell(row, tgt_key_col_idx).value
                if key_val is None or key_val == "":
                    # No key → no aggregation possible.
                    # Only fill 0 when the row actually has a SUMIFS formula string
                    # (i.e. a real data row whose key cell is empty). Rows with no
                    # formula at all (completely empty rows) must be left untouched so
                    # remove_empty_rows can still identify and delete them.
                    row_has_formula = any(
                        isinstance(ws.cell(row, tgt_col).value, str)
                        and ws.cell(row, tgt_col).value.startswith("=")
                        for tgt_col in sumifs_cols
                    )
                    if row_has_formula:
                        for tgt_col in sumifs_cols:
                            ws.cell(row, tgt_col).value = 0
                    continue

                row_sums = aggregated.get(str(key_val), {})
                for tgt_col in sumifs_cols:
                    ws.cell(row, tgt_col).value = row_sums.get(tgt_col, 0)

    # --- Phase 5: Division formulas ---
    # Scan row 2 to find division formula columns; map target_col_idx → (num_col_idx, den_col_idx)
    div_cols: dict[int, tuple[int, int]] = {}
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(2, col).value
        if not isinstance(cell_val, str):
            continue
        m = DIV_RE.match(cell_val)
        if m:
            num_col_idx = column_index_from_string(m.group(1))
            den_col_idx = column_index_from_string(m.group(2))
            div_cols[col] = (num_col_idx, den_col_idx)

    for row in range(2, ws.max_row + 1):
        for tgt_col, (num_col_idx, den_col_idx) in div_cols.items():
            num_raw = ws.cell(row, num_col_idx).value
            den_raw = ws.cell(row, den_col_idx).value
            try:
                result = float(num_raw) / float(den_raw)
            except (ZeroDivisionError, TypeError, ValueError):
                result = 0
            ws.cell(row, tgt_col).value = result

    # --- Phase 6: Residual sweep ---
    # Replace any remaining formula strings (anything starting with '=') with 0
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row, col).value
            if isinstance(val, str) and val.startswith("="):
                ws.cell(row, col).value = 0


def remove_empty_rows(ws: Worksheet) -> None:
    """Delete rows where column 1 and column 2 are both None. Row 1 (header) is never deleted.

    Also fixes a pre-existing openpyxl issue where delete_rows shifts cell values
    but leaves hyperlink.ref pointing to the old row. If left unsynced, the saved
    file's hyperlinks reference rows that no longer exist, inflating the stored
    dimension and producing trailing phantom empty rows on reload.
    """
    for row in range(ws.max_row, 1, -1):
        if ws.cell(row, 1).value is None and ws.cell(row, 2).value is None:
            ws.delete_rows(row)

    # Resync hyperlink refs with the cell's current coordinate.
    for cell in (c for col in ws.iter_cols(values_only=False) for c in col):
        if cell.hyperlink is not None and cell.hyperlink.ref != cell.coordinate:
            cell.hyperlink.ref = cell.coordinate


def process_file(src: Path, dst: Path, on_step=None) -> None:
    if on_step: on_step("正在复制文件...")
    shutil.copy2(src, dst)
    wb = load_workbook(dst, data_only=False)  # data_only=False to preserve formula strings for resolve_formulas
    ws = wb.worksheets[3]
    if on_step: on_step("正在解析公式...")
    resolve_formulas(wb, ws)
    if on_step: on_step("正在过滤空行...")
    remove_empty_rows(ws)
    if on_step: on_step("正在计算列...")
    apply_calculated_columns(ws)
    if on_step: on_step("正在排序...")
    sort_by_column(ws, "实际成本")
    if on_step: on_step("正在生成分组统计...")
    group_and_merge(ws)
    wb.save(dst)
