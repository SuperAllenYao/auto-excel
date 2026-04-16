import pytest
from openpyxl import Workbook

SHEET4_HEADERS = ["日期", "计划名称", "花费", "展现量", "点击量", "留资人数", "留资成本", "互动成本"]

FORMULA_SHEET4_HEADERS = ["笔记标题", "笔记ID", "花费", "展现量", "点击量", "留资人数", "留资成本", "互动成本"]
SOURCE_SHEET_HEADERS = ["笔记ID", "日期", "消费", "展现量", "点击量", "留资人数"]


@pytest.fixture
def make_sample_workbook():
    """Factory fixture: returns a function that creates an in-memory Workbook.

    The returned workbook has 4 sheets; Sheet 4 (index 3) contains
    marketing data with SHEET4_HEADERS in row 1 and the provided rows
    as subsequent data rows.
    """

    def _factory(rows: list[dict]) -> Workbook:
        wb = Workbook()
        # Workbook starts with 1 active sheet; add 3 more to reach 4 total.
        for name in ["Sheet2", "Sheet3", "Sheet4"]:
            wb.create_sheet(name)

        ws = wb.worksheets[3]  # Sheet4

        # Write header row
        for col_idx, header in enumerate(SHEET4_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Write data rows
        for row_idx, row_dict in enumerate(rows, start=2):
            for col_idx, header in enumerate(SHEET4_HEADERS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row_dict.get(header, 0))

        return wb

    return _factory


@pytest.fixture
def make_formula_workbook():
    """Factory fixture: returns a function that creates an in-memory Workbook
    with formula strings in Sheet 4 (index 3, title '笔记id') and a source
    data sheet (title '源数据').

    The formula sheet headers are FORMULA_SHEET4_HEADERS.
    Columns A-B contain literal values; columns C-H contain SUMIFS/division
    formula strings (not computed values).
    """

    def _factory(rows: list[dict], source_rows: list[dict]) -> Workbook:
        wb = Workbook()
        # Workbook starts with 1 active sheet ("Sheet"); add 3 more to reach
        # index 3 as the formula sheet.
        wb.create_sheet("Sheet2")
        wb.create_sheet("Sheet3")
        ws_formula = wb.create_sheet("笔记id")  # index 3

        # Write header row for formula sheet
        for col_idx, header in enumerate(FORMULA_SHEET4_HEADERS, start=1):
            ws_formula.cell(row=1, column=col_idx, value=header)

        # Write data rows with formula strings
        for row_idx, row_dict in enumerate(rows, start=2):
            # Col A: 笔记标题 (literal)
            ws_formula.cell(row=row_idx, column=1, value=row_dict.get("笔记标题", ""))
            # Col B: 笔记ID (literal)
            ws_formula.cell(row=row_idx, column=2, value=row_dict.get("笔记ID", ""))
            # Col C: 花费 — SUMIFS on 消费 (col C of 源数据)
            ws_formula.cell(
                row=row_idx,
                column=3,
                value=f"=SUMIFS('源数据'!C:C,'源数据'!A:A,B{row_idx})",
            )
            # Col D: 展现量 — SUMIFS on 展现量 (col D of 源数据)
            ws_formula.cell(
                row=row_idx,
                column=4,
                value=f"=SUMIFS('源数据'!D:D,'源数据'!A:A,B{row_idx})",
            )
            # Col E: 点击量 — SUMIFS on 点击量 (col E of 源数据)
            ws_formula.cell(
                row=row_idx,
                column=5,
                value=f"=SUMIFS('源数据'!E:E,'源数据'!A:A,B{row_idx})",
            )
            # Col F: 留资人数 — SUMIFS on 留资人数 (col F of 源数据)
            ws_formula.cell(
                row=row_idx,
                column=6,
                value=f"=SUMIFS('源数据'!F:F,'源数据'!A:A,B{row_idx})",
            )
            # Col G: 留资成本 = 花费 / 留资人数
            ws_formula.cell(row=row_idx, column=7, value=f"=C{row_idx}/F{row_idx}")
            # Col H: 互动成本 = 花费 / 点击量
            ws_formula.cell(row=row_idx, column=8, value=f"=C{row_idx}/E{row_idx}")

        # Create source data sheet
        ws_source = wb.create_sheet("源数据")  # index 4

        # Write header row for source sheet
        for col_idx, header in enumerate(SOURCE_SHEET_HEADERS, start=1):
            ws_source.cell(row=1, column=col_idx, value=header)

        # Write source data rows
        for row_idx, row_dict in enumerate(source_rows, start=2):
            for col_idx, header in enumerate(SOURCE_SHEET_HEADERS, start=1):
                ws_source.cell(row=row_idx, column=col_idx, value=row_dict.get(header, ""))

        return wb

    return _factory


@pytest.fixture
def tmp_dirs(tmp_path):
    """Fixture that creates temp Raw/, New/, and log/ subdirectories.

    Returns a dict with keys: base, raw, new, log.
    """
    raw = tmp_path / "Raw"
    new = tmp_path / "New"
    log = tmp_path / "log"

    raw.mkdir()
    new.mkdir()
    log.mkdir()

    return {"base": tmp_path, "raw": raw, "new": new, "log": log}
