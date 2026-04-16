"""Tests for the formula workbook fixture and formula resolution infrastructure."""

from conftest import FORMULA_SHEET4_HEADERS, SOURCE_SHEET_HEADERS
from auto_excel.processor import resolve_formulas


# ---------------------------------------------------------------------------
# SUMIFS resolution tests
# ---------------------------------------------------------------------------


def test_resolve_sumifs_basic(make_formula_workbook):
    """SUMIFS formulas are resolved to aggregated values from source sheet."""
    wb = make_formula_workbook(
        rows=[{"笔记标题": "Note A", "笔记ID": "id1"}, {"笔记标题": "Note B", "笔记ID": "id2"}],
        source_rows=[
            {"笔记ID": "id1", "消费": 50.0,  "展现量": 300, "点击量": 10, "留资人数": 1},
            {"笔记ID": "id1", "消费": 30.0,  "展现量": 200, "点击量": 5,  "留资人数": 2},
            {"笔记ID": "id2", "消费": 100.0, "展现量": 800, "点击量": 40, "留资人数": 3},
        ],
    )
    ws = wb.worksheets[3]
    resolve_formulas(wb, ws)

    # id1: 消费 sum = 80, 展现量 sum = 500, 点击量 sum = 15, 留资人数 sum = 3
    assert ws.cell(2, 3).value == 80.0
    assert ws.cell(2, 4).value == 500
    assert ws.cell(2, 5).value == 15
    assert ws.cell(2, 6).value == 3

    # id2 sums
    assert ws.cell(3, 3).value == 100.0
    assert ws.cell(3, 4).value == 800


def test_resolve_sumifs_missing_id(make_formula_workbook):
    wb = make_formula_workbook(
        rows=[{"笔记标题": "Missing", "笔记ID": "id_missing"}],
        source_rows=[{"笔记ID": "id_other", "消费": 100.0, "展现量": 500, "点击量": 20, "留资人数": 1}],
    )
    ws = wb.worksheets[3]
    resolve_formulas(wb, ws)
    assert ws.cell(2, 3).value == 0


def test_resolve_sumifs_no_formulas(make_sample_workbook):
    wb = make_sample_workbook([{"花费": 100.0, "展现量": 1000, "点击量": 50, "留资人数": 5}])
    assert len(wb.worksheets) >= 4  # guard against fixture shape drift
    ws = wb.worksheets[3]
    resolve_formulas(wb, ws)
    assert ws.cell(2, 3).value == 100.0  # Value unchanged


def test_resolve_division_only_without_sumifs():
    """Division formulas must still be resolved even when no SUMIFS formulas exist."""
    from openpyxl import Workbook as _Workbook
    wb = _Workbook()
    for _ in range(3):
        wb.create_sheet()
    ws = wb.worksheets[3]
    ws.cell(1, 1).value = "A"
    ws.cell(1, 2).value = "B"
    ws.cell(1, 3).value = "R"
    ws.cell(2, 1).value = 10
    ws.cell(2, 2).value = 2
    ws.cell(2, 3).value = "=A2/B2"
    resolve_formulas(wb, ws)
    assert ws.cell(2, 3).value == 5.0


def test_resolve_residual_only_without_sumifs():
    """Residual formula strings must be cleaned even when no SUMIFS formulas exist."""
    from openpyxl import Workbook as _Workbook
    wb = _Workbook()
    for _ in range(3):
        wb.create_sheet()
    ws = wb.worksheets[3]
    ws.cell(1, 1).value = "X"
    ws.cell(2, 1).value = "=IFERROR(1/0,0)"  # neither SUMIFS nor DIV pattern
    resolve_formulas(wb, ws)
    assert ws.cell(2, 1).value == 0


def test_resolve_sumifs_multiple_source_rows(make_formula_workbook):
    wb = make_formula_workbook(
        rows=[{"笔记标题": "Multi", "笔记ID": "id1"}],
        source_rows=[
            {"笔记ID": "id1", "消费": 10.0, "展现量": 100, "点击量": 5, "留资人数": 0},
            {"笔记ID": "id1", "消费": 20.0, "展现量": 200, "点击量": 3, "留资人数": 1},
            {"笔记ID": "id1", "消费": 30.0, "展现量": 300, "点击量": 7, "留资人数": 0},
        ],
    )
    ws = wb.worksheets[3]
    resolve_formulas(wb, ws)
    assert ws.cell(2, 3).value == 60.0
    assert ws.cell(2, 4).value == 600
    assert ws.cell(2, 5).value == 15
    assert ws.cell(2, 6).value == 1


def test_resolve_sumifs_source_sheet_missing_fills_all_cols(make_formula_workbook):
    """Source sheet absent → ALL SUMIFS columns filled with 0, not just one column."""
    wb = make_formula_workbook(
        rows=[{"笔记标题": "X", "笔记ID": "id1"}],
        source_rows=[{"笔记ID": "id1", "消费": 99, "展现量": 99, "点击量": 99, "留资人数": 99}],
    )
    del wb["源数据"]  # force the missing-sheet branch
    ws = wb.worksheets[3]
    resolve_formulas(wb, ws)
    assert ws.cell(2, 3).value == 0
    assert ws.cell(2, 4).value == 0
    assert ws.cell(2, 5).value == 0
    assert ws.cell(2, 6).value == 0


def test_resolve_sumifs_none_target_key_does_not_match_empty_source_key(make_formula_workbook):
    """Target row with key=None must NOT aggregate source rows whose key is ''."""
    wb = make_formula_workbook(
        rows=[{"笔记标题": "No ID"}],
        source_rows=[{"笔记ID": "", "消费": 999.0, "展现量": 0, "点击量": 0, "留资人数": 0}],
    )
    ws = wb.worksheets[3]
    ws.cell(2, 2).value = None  # explicitly wipe the target key cell
    resolve_formulas(wb, ws)
    assert ws.cell(2, 3).value == 0


def test_resolve_sumifs_each_col_aggregated_independently(make_formula_workbook):
    """Each SUMIFS target column must sum its own source column, not share a total."""
    wb = make_formula_workbook(
        rows=[{"笔记标题": "T", "笔记ID": "id1"}],
        source_rows=[
            {"笔记ID": "id1", "消费": 10.0, "展现量": 100, "点击量": 5, "留资人数": 1},
            {"笔记ID": "id1", "消费": 20.0, "展现量": 200, "点击量": 10, "留资人数": 2},
        ],
    )
    ws = wb.worksheets[3]
    resolve_formulas(wb, ws)
    assert ws.cell(2, 3).value == 30.0
    assert ws.cell(2, 4).value == 300
    assert ws.cell(2, 5).value == 15
    assert ws.cell(2, 6).value == 3
    # Guards against an impl that computes one total and broadcasts to every col.
    assert ws.cell(2, 3).value != ws.cell(2, 4).value


def test_resolve_division_multi_row_distinct_values():
    """Each row must divide with its own numerator/denominator, not row 2's."""
    from openpyxl import Workbook as _Workbook
    wb = _Workbook()
    for _ in range(3):
        wb.create_sheet()
    ws = wb.worksheets[3]
    ws.cell(1, 1).value = "Num"
    ws.cell(1, 2).value = "Den"
    ws.cell(1, 3).value = "Ratio"
    # Row 2: 10 / 2 = 5.0
    ws.cell(2, 1).value = 10
    ws.cell(2, 2).value = 2
    ws.cell(2, 3).value = "=A2/B2"
    # Row 3: 30 / 3 = 10.0 (distinct from row 2 to catch broadcast)
    ws.cell(3, 1).value = 30
    ws.cell(3, 2).value = 3
    ws.cell(3, 3).value = "=A3/B3"
    # Row 4: 7 / 2 = 3.5
    ws.cell(4, 1).value = 7
    ws.cell(4, 2).value = 2
    ws.cell(4, 3).value = "=A4/B4"
    resolve_formulas(wb, ws)
    assert ws.cell(2, 3).value == 5.0
    assert ws.cell(3, 3).value == 10.0  # must NOT be row-2's 5.0
    assert ws.cell(4, 3).value == 3.5


def test_resolve_phase6_does_not_erase_phase5_results():
    """Phase 6 residual sweep must leave Phase 5's float results intact."""
    from openpyxl import Workbook as _Workbook
    wb = _Workbook()
    for _ in range(3):
        wb.create_sheet()
    ws = wb.worksheets[3]
    ws.cell(1, 1).value = "Num"
    ws.cell(1, 2).value = "Den"
    ws.cell(1, 3).value = "Ratio"
    ws.cell(1, 4).value = "Residual"
    ws.cell(2, 1).value = 12
    ws.cell(2, 2).value = 4
    ws.cell(2, 3).value = "=A2/B2"          # Phase 5 → 3.0
    ws.cell(2, 4).value = "=IFERROR(A2,0)"  # Phase 6 cleans this
    resolve_formulas(wb, ws)
    assert ws.cell(2, 3).value == 3.0  # Phase 5 result must survive
    assert ws.cell(2, 4).value == 0


def test_resolve_sumifs_then_division_multi_row_distinct(make_formula_workbook):
    """Multi-row SUMIFS→Division chain: each row's division must use its own sums."""
    wb = make_formula_workbook(
        rows=[
            {"笔记标题": "A", "笔记ID": "id1"},
            {"笔记标题": "B", "笔记ID": "id2"},
        ],
        source_rows=[
            {"笔记ID": "id1", "消费": 60.0, "展现量": 300, "点击量": 6, "留资人数": 3},
            {"笔记ID": "id2", "消费": 20.0, "展现量": 100, "点击量": 4, "留资人数": 2},
        ],
    )
    ws = wb.worksheets[3]
    resolve_formulas(wb, ws)
    # Row 2 (id1): 留资成本 = 60/3 = 20.0; 互动成本 = 60/6 = 10.0
    assert ws.cell(2, 7).value == 20.0
    assert ws.cell(2, 8).value == 10.0
    # Row 3 (id2): 留资成本 = 20/2 = 10.0; 互动成本 = 20/4 = 5.0
    assert ws.cell(3, 7).value == 10.0
    assert ws.cell(3, 8).value == 5.0
    # Guard against an impl that broadcasts row 2's result to row 3.
    assert ws.cell(3, 7).value != ws.cell(2, 7).value
    assert ws.cell(3, 8).value != ws.cell(2, 8).value


# ---------------------------------------------------------------------------
# Smoke test: fixture creates correct structure
# ---------------------------------------------------------------------------


def test_fixture_creates_formula_workbook(make_formula_workbook):
    wb = make_formula_workbook(
        rows=[{"笔记标题": "Test", "笔记ID": "id1"}],
        source_rows=[
            {
                "笔记ID": "id1",
                "日期": "2026-01-01",
                "消费": 100.0,
                "展现量": 500,
                "点击量": 20,
                "留资人数": 2,
            }
        ],
    )
    ws = wb.worksheets[3]
    assert ws.cell(1, 3).value == "花费"
    assert isinstance(ws.cell(2, 3).value, str)
    assert ws.cell(2, 3).value.startswith("=SUMIFS")


# ---------------------------------------------------------------------------
# Adversarial tests
# ---------------------------------------------------------------------------


# Category 1: Wrong data — missing 笔记ID key should not crash; formula
# row reference must still be correct.
def test_fixture_missing_notebook_id_defaults_to_empty(make_formula_workbook):
    """If a row dict omits 笔记ID, col B should be empty string (not raise)."""
    wb = make_formula_workbook(
        rows=[{"笔记标题": "Only title"}],
        source_rows=[],
    )
    ws = wb.worksheets[3]
    assert ws.cell(2, 2).value == ""


# Category 2: Structural integrity — workbook must have 5 sheets in the right
# order and with the right titles.
def test_fixture_sheet_structure(make_formula_workbook):
    """Formula sheet must be at index 3 with title '笔记id'; source at index 4 '源数据'."""
    wb = make_formula_workbook(rows=[], source_rows=[])
    assert len(wb.worksheets) == 5
    assert wb.worksheets[3].title == "笔记id"
    assert wb.worksheets[4].title == "源数据"


# Category 3: Formula string correctness — each formula column should
# reference the correct source column and the right row number.
def test_fixture_formula_strings_reference_correct_columns(make_formula_workbook):
    """All eight columns of row 2 must have the values/formulas described in the spec."""
    wb = make_formula_workbook(
        rows=[{"笔记标题": "Note A", "笔记ID": "abc"}],
        source_rows=[],
    )
    ws = wb.worksheets[3]

    # Literal columns
    assert ws.cell(2, 1).value == "Note A"
    assert ws.cell(2, 2).value == "abc"

    # SUMIFS formulas — verify both source column letter and row reference
    assert ws.cell(2, 3).value == "=SUMIFS('源数据'!C:C,'源数据'!A:A,B2)"  # 花费
    assert ws.cell(2, 4).value == "=SUMIFS('源数据'!D:D,'源数据'!A:A,B2)"  # 展现量
    assert ws.cell(2, 5).value == "=SUMIFS('源数据'!E:E,'源数据'!A:A,B2)"  # 点击量
    assert ws.cell(2, 6).value == "=SUMIFS('源数据'!F:F,'源数据'!A:A,B2)"  # 留资人数

    # Division formulas
    assert ws.cell(2, 7).value == "=C2/F2"   # 留资成本
    assert ws.cell(2, 8).value == "=C2/E2"   # 互动成本


# Category 4: Multiple rows — row index in formulas must advance correctly.
def test_fixture_multi_row_formula_row_references(make_formula_workbook):
    """Row 3 formulas must reference row 3, not row 2."""
    wb = make_formula_workbook(
        rows=[
            {"笔记标题": "A", "笔记ID": "id1"},
            {"笔记标题": "B", "笔记ID": "id2"},
        ],
        source_rows=[],
    )
    ws = wb.worksheets[3]
    assert ws.cell(3, 3).value == "=SUMIFS('源数据'!C:C,'源数据'!A:A,B3)"
    assert ws.cell(3, 7).value == "=C3/F3"
    assert ws.cell(3, 8).value == "=C3/E3"


# Category 5: Source sheet — headers and data rows must be written correctly.
def test_fixture_source_sheet_headers_and_data(make_formula_workbook):
    """Source sheet must have correct headers and data rows."""
    wb = make_formula_workbook(
        rows=[],
        source_rows=[
            {
                "笔记ID": "note01",
                "日期": "2026-03-01",
                "消费": 250.5,
                "展现量": 1000,
                "点击量": 50,
                "留资人数": 5,
            }
        ],
    )
    ws_src = wb.worksheets[4]

    # Header row
    for col_idx, expected_header in enumerate(SOURCE_SHEET_HEADERS, start=1):
        assert ws_src.cell(1, col_idx).value == expected_header

    # Data row — assert every column so swapped writes are caught.
    assert ws_src.cell(2, 1).value == "note01"
    assert ws_src.cell(2, 2).value == "2026-03-01"
    assert ws_src.cell(2, 3).value == 250.5
    assert ws_src.cell(2, 4).value == 1000
    assert ws_src.cell(2, 5).value == 50
    assert ws_src.cell(2, 6).value == 5


# Category 6: Header row of formula sheet must match FORMULA_SHEET4_HEADERS exactly.
def test_fixture_formula_sheet_headers(make_formula_workbook):
    """Formula sheet row 1 must exactly match FORMULA_SHEET4_HEADERS."""
    wb = make_formula_workbook(rows=[], source_rows=[])
    ws = wb.worksheets[3]
    # Read 9 cols (one past expected end) to also verify there are no extra headers.
    actual_headers = [ws.cell(1, col).value for col in range(1, 10)]
    assert actual_headers[:8] == FORMULA_SHEET4_HEADERS
    assert actual_headers[8] is None  # no stray header at col 9


# Category 7: Constant integrity — guard against the constant itself being truncated.
def test_formula_sheet4_headers_constant_integrity():
    """FORMULA_SHEET4_HEADERS must have exactly 8 items with known first and last."""
    assert len(FORMULA_SHEET4_HEADERS) == 8
    assert FORMULA_SHEET4_HEADERS[0] == "笔记标题"
    assert FORMULA_SHEET4_HEADERS[7] == "互动成本"


# Category 8: Numeric default behavior — missing numeric keys in source_rows default to 0.
def test_fixture_source_numeric_default_is_zero(make_formula_workbook):
    """Missing numeric keys in source_rows default to 0 so float() is safe downstream."""
    wb = make_formula_workbook(
        rows=[],
        source_rows=[{"笔记ID": "id9", "日期": "2026-01-01"}],  # numeric cols omitted
    )
    ws_src = wb.worksheets[4]
    assert ws_src.cell(2, 3).value == 0  # 消费
    assert ws_src.cell(2, 4).value == 0  # 展现量
    assert ws_src.cell(2, 5).value == 0  # 点击量
    assert ws_src.cell(2, 6).value == 0  # 留资人数


# ---------------------------------------------------------------------------
# Division formula resolution tests
# ---------------------------------------------------------------------------


def test_resolve_division_formulas(make_formula_workbook):
    """Col G (留资成本) = 花费 / 留资人数 is computed after SUMIFS; cell must be numeric."""
    wb = make_formula_workbook(
        rows=[{"笔记标题": "Note A", "笔记ID": "id1"}],
        source_rows=[
            {"笔记ID": "id1", "消费": 80.0, "展现量": 500, "点击量": 20, "留资人数": 4},
        ],
    )
    ws = wb.worksheets[3]
    resolve_formulas(wb, ws)

    # Col C (花费) should be 80.0, Col F (留资人数) should be 4
    # Col G (留资成本) = 80 / 4 = 20.0 — must be numeric, not a formula string
    result = ws.cell(2, 7).value
    assert isinstance(result, (int, float)), f"Expected numeric, got {type(result)}: {result!r}"
    assert abs(result - 20.0) < 1e-9


def test_resolve_division_by_zero(make_formula_workbook):
    """When denominator resolves to 0, division result must be 0 (not error, not formula string)."""
    wb = make_formula_workbook(
        rows=[{"笔记标题": "Note A", "笔记ID": "id1"}],
        source_rows=[
            {"笔记ID": "id1", "消费": 80.0, "展现量": 500, "点击量": 20, "留资人数": 0},
        ],
    )
    ws = wb.worksheets[3]
    resolve_formulas(wb, ws)

    # 留资人数 = 0 → 留资成本 (col G) = 0
    result = ws.cell(2, 7).value
    assert result == 0, f"Expected 0 for division by zero, got {result!r}"
    # 点击量 = 20, so 互动成本 (col H) = 80 / 20 = 4.0
    assert abs(ws.cell(2, 8).value - 4.0) < 1e-9


def test_resolve_residual_formula_strings(make_formula_workbook):
    """Any '=' formula string remaining after SUMIFS and division resolution is cleaned to 0."""
    wb = make_formula_workbook(
        rows=[{"笔记标题": "Note A", "笔记ID": "id1"}],
        source_rows=[
            {"笔记ID": "id1", "消费": 10.0, "展现量": 100, "点击量": 5, "留资人数": 1},
        ],
    )
    ws = wb.worksheets[3]
    # Plant an unhandled formula that matches neither SUMIFS_RE nor DIV_RE
    ws.cell(2, 8).value = "=IFERROR(C2/E2,0)"  # not matched by DIV_RE
    resolve_formulas(wb, ws)

    # Residual sweep must replace any remaining '=' strings with 0
    result = ws.cell(2, 8).value
    assert result == 0, f"Expected 0 for residual formula, got {result!r}"


def test_adversarial_division_depends_on_sumifs(make_formula_workbook):
    """Chained: SUMIFS resolves 花费=80, 留资人数=3 → 留资成本=80/3 ≈ 26.67."""
    wb = make_formula_workbook(
        rows=[{"笔记标题": "Chain", "笔记ID": "id1"}],
        source_rows=[
            {"笔记ID": "id1", "消费": 50.0, "展现量": 200, "点击量": 10, "留资人数": 2},
            {"笔记ID": "id1", "消费": 30.0, "展现量": 100, "点击量": 5,  "留资人数": 1},
        ],
    )
    ws = wb.worksheets[3]
    resolve_formulas(wb, ws)

    # After SUMIFS: 花费 = 80 (col C), 留资人数 = 3 (col F)
    assert ws.cell(2, 3).value == 80.0
    assert ws.cell(2, 6).value == 3.0

    # Division must use the SUMIFS-resolved values, not formula strings
    result = ws.cell(2, 7).value
    assert isinstance(result, (int, float)), f"Expected numeric, got {type(result)}: {result!r}"
    expected = 80.0 / 3.0
    assert abs(result - expected) < 1e-9, f"Expected {expected}, got {result}"


def test_adversarial_all_zero_source_data(make_formula_workbook):
    """All source data zeros → all SUMIFS=0 → all divisions=0 (no errors)."""
    wb = make_formula_workbook(
        rows=[
            {"笔记标题": "Row1", "笔记ID": "id1"},
            {"笔记标题": "Row2", "笔记ID": "id2"},
        ],
        source_rows=[
            {"笔记ID": "id1", "消费": 0, "展现量": 0, "点击量": 0, "留资人数": 0},
            {"笔记ID": "id2", "消费": 0, "展现量": 0, "点击量": 0, "留资人数": 0},
        ],
    )
    ws = wb.worksheets[3]
    resolve_formulas(wb, ws)

    for row in (2, 3):
        for col in range(3, 9):  # cols C through H
            val = ws.cell(row, col).value
            assert val == 0, f"Expected 0 at row={row} col={col}, got {val!r}"
            assert not isinstance(val, str), f"Got formula string at row={row} col={col}: {val!r}"
