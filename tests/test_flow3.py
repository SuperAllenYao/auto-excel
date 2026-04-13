from auto_excel.processor import apply_calculated_columns, sort_by_column, group_and_merge, find_column

# Helper: 5-row workbook where 实际成本 will be: 100, 100, ~66.7, 25, 20
ROWS_5 = [
    {"花费": 113.6, "展现量": 1000, "点击量": 50, "留资人数": 1,   "留资成本": 100, "互动成本": 10},
    {"花费": 113.6, "展现量": 1000, "点击量": 50, "留资人数": 1,   "留资成本": 100, "互动成本": 10},
    {"花费": 113.6, "展现量": 1000, "点击量": 50, "留资人数": 1.5, "留资成本": 67,  "互动成本": 10},
    {"花费": 113.6, "展现量": 1000, "点击量": 50, "留资人数": 4,   "留资成本": 25,  "互动成本": 10},
    {"花费": 113.6, "展现量": 1000, "点击量": 50, "留资人数": 5,   "留资成本": 20,  "互动成本": 10},
]

def _apply_all(ws):
    apply_calculated_columns(ws)
    sort_by_column(ws, "实际成本", descending=True)
    group_and_merge(ws)

def test_group_inserts_zhanbi_column(make_sample_workbook):
    wb = make_sample_workbook(ROWS_5)
    ws = wb.worksheets[3]
    _apply_all(ws)
    assert find_column(ws, "占比") > 0

def test_group_merge_content(make_sample_workbook):
    wb = make_sample_workbook(ROWS_5)
    ws = wb.worksheets[3]
    _apply_all(ws)
    zb_col = find_column(ws, "占比")
    # After sort desc: rows 2-3 are ≥90, row 4 is 50-89, rows 5-6 are <50
    assert ws.cell(2, zb_col).value == "2/40%"
    assert ws.cell(4, zb_col).value == "1/20%"
    assert ws.cell(5, zb_col).value == "2/40%"

def test_adversarial_all_in_one_range(make_sample_workbook):
    """All rows ≥ 90 → single group '5/100%', no other groups."""
    rows = [
        {"花费": 113.6, "展现量": 1000, "点击量": 50, "留资人数": 1, "留资成本": 100, "互动成本": 10},
    ] * 5
    wb = make_sample_workbook(rows)
    ws = wb.worksheets[3]
    _apply_all(ws)
    zb_col = find_column(ws, "占比")
    assert ws.cell(2, zb_col).value == "5/100%"
    # Rows 3-6 are merged cells — their value should be None (not "5/100%")
    assert ws.cell(3, zb_col).value is None

def test_adversarial_empty_range(make_sample_workbook):
    """No medium range rows → only high and low groups appear."""
    rows = [
        {"花费": 113.6, "展现量": 1000, "点击量": 50, "留资人数": 1, "留资成本": 100, "互动成本": 10},
        {"花费": 113.6, "展现量": 1000, "点击量": 50, "留资人数": 1, "留资成本": 100, "互动成本": 10},
        {"花费": 113.6, "展现量": 1000, "点击量": 50, "留资人数": 5, "留资成本": 20,  "互动成本": 10},
        {"花费": 113.6, "展现量": 1000, "点击量": 50, "留资人数": 4, "留资成本": 20,  "互动成本": 10},
    ]
    wb = make_sample_workbook(rows)
    ws = wb.worksheets[3]
    _apply_all(ws)
    zb_col = find_column(ws, "占比")
    # Rows 2-3: high (≥90), 2 rows, 50%
    assert ws.cell(2, zb_col).value == "2/50%"
    # Rows 4-5: low (<50), 2 rows, 50%
    assert ws.cell(4, zb_col).value == "2/50%"

def test_adversarial_single_row(make_sample_workbook):
    """Single data row → '1/100%', no merge attempted."""
    wb = make_sample_workbook([
        {"花费": 113.6, "展现量": 1000, "点击量": 50, "留资人数": 1, "留资成本": 100, "互动成本": 10}
    ])
    ws = wb.worksheets[3]
    _apply_all(ws)
    zb_col = find_column(ws, "占比")
    assert ws.cell(2, zb_col).value == "1/100%"

def test_adversarial_zhanbi_column_position(make_sample_workbook):
    """占比列必须紧跟互动成本列之后，不能插在末尾或其他位置。"""
    wb = make_sample_workbook(ROWS_5)
    ws = wb.worksheets[3]
    apply_calculated_columns(ws)
    sort_by_column(ws, "实际成本", descending=True)
    hd_col_before_merge = find_column(ws, "互动成本")
    group_and_merge(ws)
    hd_col_after = find_column(ws, "互动成本")
    zb_col = find_column(ws, "占比")
    assert zb_col == hd_col_after + 1

def test_adversarial_boundary_values_90_and_50():
    """实际成本 = 90 必须归入 high（≥90），= 50 必须归入 medium（≥50）。"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    # Minimal sheet with just 互动成本 and 实际成本 columns
    ws.cell(1, 1).value = "互动成本"
    ws.cell(1, 2).value = "实际成本"
    ws.cell(2, 1).value = 10;  ws.cell(2, 2).value = 90   # boundary: exactly 90 → high
    ws.cell(3, 1).value = 10;  ws.cell(3, 2).value = 50   # boundary: exactly 50 → medium
    ws.cell(4, 1).value = 10;  ws.cell(4, 2).value = 20   # low

    group_and_merge(ws)
    zb_col = find_column(ws, "占比")

    # 90 → high (1 row, 33% of 3)
    assert ws.cell(2, zb_col).value == "1/33%", f"实际成本=90 应属于 high 组，got {ws.cell(2, zb_col).value}"
    # 50 → medium (1 row, 33% of 3)
    assert ws.cell(3, zb_col).value == "1/33%", f"实际成本=50 应属于 medium 组，got {ws.cell(3, zb_col).value}"
    # 20 → low (1 row, 33% of 3)
    assert ws.cell(4, zb_col).value == "1/33%"
