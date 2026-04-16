"""End-to-end integration test for the full auto-excel pipeline."""
import json
from openpyxl import load_workbook
from typer.testing import CliRunner
from auto_excel.cli import app

runner = CliRunner()

ROWS = [
    {"花费": 100.0,  "展现量": 3000,  "点击量": 100, "留资人数": 4},   # 实际成本≈22.0  (低)
    {"花费": 450.0,  "展现量": 8000,  "点击量": 400, "留资人数": 4},   # 实际成本≈99.0  (高)
    {"花费": 700.0,  "展现量": 12000, "点击量": 600, "留资人数": 5},   # 实际成本≈123.2 (高)
    {"花费": 200.0,  "展现量": 5000,  "点击量": 200, "留资人数": 3},   # 实际成本≈58.7  (中)
    {"花费": 150.0,  "展现量": 4000,  "点击量": 150, "留资人数": 5},   # 实际成本≈26.4  (低)
    {"花费": 600.0,  "展现量": 10000, "点击量": 500, "留资人数": 5},   # 实际成本≈105.6 (高)
]


def test_full_pipeline(tmp_dirs, monkeypatch, make_sample_workbook):
    """Full pipeline: Raw/ → process → New/ with all transformations verified."""
    import auto_excel.config as cfg
    import auto_excel.state as st

    state_file = tmp_dirs["log"] / "processed.json"
    for attr, val in [("RAW_DIR", tmp_dirs["raw"]), ("NEW_DIR", tmp_dirs["new"]),
                      ("LOG_DIR", tmp_dirs["log"]), ("STATE_FILE", state_file)]:
        monkeypatch.setattr(cfg, attr, val)
    monkeypatch.setattr(st, "STATE_FILE", state_file)

    # 1. Save sample workbook to Raw/
    wb = make_sample_workbook(ROWS)
    src = tmp_dirs["raw"] / "integration_test.xlsx"
    wb.save(src)

    # 2. Run CLI
    result = runner.invoke(app, ["on"])
    assert result.exit_code == 0, f"CLI failed: {result.output}"

    # 3. Output file exists in New/
    out_path = tmp_dirs["new"] / "integration_test.xlsx"
    assert out_path.exists()

    # 4. State recorded
    assert state_file.exists()
    state = json.loads(state_file.read_text())
    assert "integration_test.xlsx" in state

    # 5. Open output and check Sheet 4
    wb_out = load_workbook(out_path)
    ws = wb_out.worksheets[3]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # 6. Verify all 4 calculated columns exist (Flow 1)
    for col in ["实际花费", "点击率", "CPC", "实际成本"]:
        assert col in headers, f"Missing column: {col}"

    # 7. Verify 占比 column exists (Flow 3)
    assert "占比" in headers, "Missing 占比 column"

    # 8. Verify data is sorted descending by 实际成本 (Flow 2)
    cost_col = headers.index("实际成本") + 1
    costs = [ws.cell(r, cost_col).value for r in range(2, ws.max_row + 1)
             if ws.cell(r, cost_col).value is not None]
    assert len(costs) == 6
    assert costs == sorted(costs, reverse=True), f"Not sorted desc: {costs}"

    # 9. Verify 实际花费 values (spot check first row after sort)
    shf_col = headers.index("实际花费") + 1
    first_shf = ws.cell(2, shf_col).value
    # First row after sort has highest cost: 花费=700/1.136 ≈ 616.197
    assert first_shf is not None
    assert abs(float(first_shf) - 700.0 / 1.136) < 0.01, f"实际花費 wrong: {first_shf}"

    # 10. Verify 占比 content with exact values for all 3 groups
    zb_col = headers.index("占比") + 1
    # High group (rows 2-4): top-left has value, others None (merged)
    assert ws.cell(2, zb_col).value == "3/50%",  f"High group 占比 wrong: {ws.cell(2, zb_col).value}"
    assert ws.cell(3, zb_col).value is None,     "Row 3 should be merged (None)"
    assert ws.cell(4, zb_col).value is None,     "Row 4 should be merged (None)"
    # Medium group (row 5): single cell
    assert ws.cell(5, zb_col).value == "1/17%",  f"Medium group 占比 wrong: {ws.cell(5, zb_col).value}"
    # Low group (rows 6-7): top-left has value, row 7 None (merged)
    assert ws.cell(6, zb_col).value == "2/33%",  f"Low group 占比 wrong: {ws.cell(6, zb_col).value}"
    assert ws.cell(7, zb_col).value is None,     "Row 7 should be merged (None)"

    # Verify each group's 实际成本 values are in the correct threshold range
    cost_col = headers.index("实际成本") + 1
    # High group (rows 2-4): all must be >= 90
    for r in range(2, 5):
        c = ws.cell(r, cost_col).value
        assert c is not None and float(c) >= 90, f"Row {r} in High group has cost {c} < 90"
    # Medium group (row 5): must be in [50, 90)
    c5 = ws.cell(5, cost_col).value
    assert c5 is not None and 50 <= float(c5) < 90, f"Row 5 in Medium group has cost {c5} outside [50,90)"
    # Low group (rows 6-7): all must be < 50
    for r in range(6, 8):
        c = ws.cell(r, cost_col).value
        assert c is not None and float(c) < 50, f"Row {r} in Low group has cost {c} >= 50"


def test_full_pipeline_with_formulas(tmp_dirs, monkeypatch, make_formula_workbook):
    """Full pipeline with formula-based data produces correct numeric output."""
    import auto_excel.config as cfg
    import auto_excel.state as st

    state_file = tmp_dirs["log"] / "processed.json"
    for attr, val in [("RAW_DIR", tmp_dirs["raw"]), ("NEW_DIR", tmp_dirs["new"]),
                      ("LOG_DIR", tmp_dirs["log"]), ("STATE_FILE", state_file)]:
        monkeypatch.setattr(cfg, attr, val)
    monkeypatch.setattr(st, "STATE_FILE", state_file)

    wb = make_formula_workbook(
        rows=[
            {"笔记标题": "High Cost Note",   "笔记ID": "id1"},
            {"笔记标题": "Medium Cost Note", "笔记ID": "id2"},
            {"笔记标题": "Low Cost Note",    "笔记ID": "id3"},
        ],
        source_rows=[
            {"笔记ID": "id1", "消费": 500.0, "展现量": 5000, "点击量": 200, "留资人数": 2},
            {"笔记ID": "id1", "消费": 300.0, "展现量": 3000, "点击量": 100, "留资人数": 1},
            {"笔记ID": "id2", "消费": 200.0, "展现量": 2000, "点击量": 80,  "留资人数": 3},
            {"笔记ID": "id3", "消费": 50.0,  "展现量": 1000, "点击量": 30,  "留资人数": 5},
        ],
    )
    src = tmp_dirs["raw"] / "formula_test.xlsx"
    wb.save(src)

    result = runner.invoke(app, ["on"])
    assert result.exit_code == 0, f"CLI failed: {result.output}"

    out_path = tmp_dirs["new"] / "formula_test.xlsx"
    assert out_path.exists()

    wb_out = load_workbook(out_path)
    ws = wb_out.worksheets[3]

    # Verify formula columns are resolved to numbers, not formula strings or None
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    huafei_col = headers.index("花费") + 1

    # After sort by 实际成本 desc: id1 has highest cost (花费 800 / 留资人数 3 ≈ 266.7),
    # id2 next (200 / 3 ≈ 66.7), id3 last (50 / 5 = 10.0). So row 2 = id1, 花费 = 800.
    first_huafei = ws.cell(2, huafei_col).value
    assert first_huafei == 800.0, \
        f"Row 2 (id1) 花费 should be exactly 800.0, got {first_huafei!r} — SUMIFS resolution wrong"

    # All data rows must have numeric 花费 > 0 (花费 column is not part of group_and_merge,
    # so no merged-None cells — any None here is a formula-resolution bug).
    data_rows = ws.max_row - 1  # subtract header
    assert data_rows == 3, f"Expected 3 data rows (one per ID), got {data_rows}"
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, huafei_col).value
        assert val is not None and isinstance(val, (int, float)) and float(val) > 0, \
            f"Row {r} 花费 should be numeric > 0, got {val!r}"

    # 占比 column must exist (group_and_merge ran)
    assert "占比" in headers, "Missing 占比 column — group_and_merge did not run"

    # 实际成本 column must exist and be sorted descending
    assert "实际成本" in headers, "Missing 实际成本 column"
    cost_col = headers.index("实际成本") + 1
    costs = [ws.cell(r, cost_col).value for r in range(2, ws.max_row + 1)
             if ws.cell(r, cost_col).value is not None]
    assert costs == sorted(costs, reverse=True), f"实际成本 not sorted desc: {costs}"


def test_pipeline_runs_resolve_formulas_before_apply_calculated_columns(
    tmp_dirs, monkeypatch, make_formula_workbook
):
    """Probe the pipeline order: resolve_formulas must run before apply_calculated_columns.

    Construct a row where 笔记ID is empty string. Its formula cells (col C-H) still contain
    formula STRINGS after load. For the pipeline to succeed:
      1. resolve_formulas must run first — it replaces formula strings in empty-key rows with 0
      2. remove_empty_rows keeps the row (col A has a title, not None)
      3. apply_calculated_columns then sees numeric 0s, not formula strings
    If resolve_formulas runs AFTER apply_calculated_columns, the latter would call float() on a
    formula string and crash or return 0. This test catches that wiring inversion.
    """
    import auto_excel.config as cfg
    import auto_excel.state as st

    state_file = tmp_dirs["log"] / "processed.json"
    for attr, val in [("RAW_DIR", tmp_dirs["raw"]), ("NEW_DIR", tmp_dirs["new"]),
                      ("LOG_DIR", tmp_dirs["log"]), ("STATE_FILE", state_file)]:
        monkeypatch.setattr(cfg, attr, val)
    monkeypatch.setattr(st, "STATE_FILE", state_file)

    wb = make_formula_workbook(
        rows=[
            {"笔记标题": "Normal", "笔记ID": "id1"},
            {"笔记标题": "Empty Key Row", "笔记ID": ""},  # formulas present, key empty
        ],
        source_rows=[
            {"笔记ID": "id1", "消费": 500.0, "展现量": 5000, "点击量": 200, "留资人数": 5},
        ],
    )
    src = tmp_dirs["raw"] / "order_probe.xlsx"
    wb.save(src)

    result = runner.invoke(app, ["on"])
    assert result.exit_code == 0, f"CLI failed: {result.output}"

    wb_out = load_workbook(tmp_dirs["new"] / "order_probe.xlsx")
    ws = wb_out.worksheets[3]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    huafei_col = headers.index("花费") + 1

    # Every data row's 花费 column must be numeric — any formula string remaining would prove
    # resolve_formulas did not run before apply_calculated_columns.
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, huafei_col).value
        assert isinstance(val, (int, float)), \
            f"Row {r} 花费={val!r} — formula not resolved, pipeline order inversion"


def test_full_pipeline_with_formulas_and_empty_rows(tmp_dirs, monkeypatch, make_formula_workbook):
    """Formula workbook with empty rows: pipeline filters empties and computes groups correctly."""
    import auto_excel.config as cfg
    import auto_excel.state as st

    state_file = tmp_dirs["log"] / "processed.json"
    for attr, val in [("RAW_DIR", tmp_dirs["raw"]), ("NEW_DIR", tmp_dirs["new"]),
                      ("LOG_DIR", tmp_dirs["log"]), ("STATE_FILE", state_file)]:
        monkeypatch.setattr(cfg, attr, val)
    monkeypatch.setattr(st, "STATE_FILE", state_file)

    wb = make_formula_workbook(
        rows=[
            {"笔记标题": "High Cost Note",   "笔记ID": "id1"},
            {"笔记标题": "Medium Cost Note", "笔记ID": "id2"},
            {"笔记标题": "Low Cost Note",    "笔记ID": "id3"},
        ],
        source_rows=[
            {"笔记ID": "id1", "消费": 800.0, "展现量": 8000, "点击量": 400, "留资人数": 4},
            {"笔记ID": "id2", "消费": 200.0, "展现量": 2000, "点击量": 80,  "留资人数": 3},
            {"笔记ID": "id3", "消费": 50.0,  "展现量": 1000, "点击量": 30,  "留资人数": 5},
        ],
    )

    # Manually append 2 empty rows to Sheet 4 (index 3) after the data rows
    ws_formula = wb.worksheets[3]
    empty_row_start = ws_formula.max_row + 1
    for empty_row in range(empty_row_start, empty_row_start + 2):
        # Leave col A and col B as None (both required for remove_empty_rows to delete)
        ws_formula.cell(row=empty_row, column=1).value = None
        ws_formula.cell(row=empty_row, column=2).value = None

    src = tmp_dirs["raw"] / "empty_rows_test.xlsx"
    wb.save(src)

    result = runner.invoke(app, ["on"])
    assert result.exit_code == 0, f"CLI failed: {result.output}"

    out_path = tmp_dirs["new"] / "empty_rows_test.xlsx"
    assert out_path.exists()

    wb_out = load_workbook(out_path)
    ws = wb_out.worksheets[3]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # Output should have exactly 3 data rows (empty rows removed)
    data_rows = ws.max_row - 1
    assert data_rows == 3, \
        f"Expected 3 data rows after empty row removal, got {data_rows} — empty rows not filtered"

    # 花費 must be numeric for all data rows (formulas resolved)
    huafei_col = headers.index("花费") + 1
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, huafei_col).value
        if val is not None:
            assert isinstance(val, (int, float)), \
                f"Row {r} 花費 not numeric after empty row removal: {val!r}"

    # 占比 percentages must total to ~100% (groups based on 3 data rows, not 5)
    assert "占比" in headers, "Missing 占比 column"
    zb_col = headers.index("占比") + 1
    zb_values = [ws.cell(r, zb_col).value for r in range(2, ws.max_row + 1)
                 if ws.cell(r, zb_col).value is not None]
    # Each 占比 is like "1/33%" or "2/67%"; extract percentages and verify sum ~100%
    total_pct = 0
    for zb in zb_values:
        # Format is "N/P%" — extract P
        pct_str = zb.split("/")[1].rstrip("%")
        total_pct += int(pct_str)
    assert abs(total_pct - 100) <= 2, \
        f"占比 percentages sum to {total_pct}%, expected ~100% for 3 data rows (not 5)"
