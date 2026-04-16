from openpyxl import Workbook
from auto_excel.processor import remove_empty_rows


def test_remove_empty_rows_basic():
    wb = Workbook()
    ws = wb.active
    ws.append(["标题", "ID", "数据"])      # row 1: header
    ws.append(["Note1", "id1", 100])        # row 2: data
    ws.append([None, None, None])           # row 3: empty
    ws.append(["Note2", "id2", 200])        # row 4: data
    ws.append([None, None, None])           # row 5: empty
    remove_empty_rows(ws)
    assert ws.max_row == 3
    assert ws.cell(2, 1).value == "Note1"
    assert ws.cell(3, 1).value == "Note2"


def test_remove_empty_rows_no_empty():
    wb = Workbook()
    ws = wb.active
    ws.append(["标题", "ID"])
    ws.append(["A", "id1"])
    ws.append(["B", "id2"])
    remove_empty_rows(ws)
    assert ws.max_row == 3


def test_remove_empty_rows_all_empty():
    wb = Workbook()
    ws = wb.active
    ws.append(["标题", "ID"])
    ws.append([None, None])
    ws.append([None, None])
    remove_empty_rows(ws)
    assert ws.max_row == 1


def test_adversarial_title_but_no_id():
    wb = Workbook()
    ws = wb.active
    ws.append(["标题", "ID"])
    ws.append(["Has Title", None])
    remove_empty_rows(ws)
    assert ws.max_row == 2
    assert ws.cell(2, 1).value == "Has Title"


def test_adversarial_id_but_no_title():
    wb = Workbook()
    ws = wb.active
    ws.append(["标题", "ID"])
    ws.append([None, "has_id"])
    remove_empty_rows(ws)
    assert ws.max_row == 2
    assert ws.cell(2, 2).value == "has_id"


def test_adversarial_empty_rows_scattered():
    wb = Workbook()
    ws = wb.active
    ws.append(["标题", "ID"])
    ws.append([None, None])
    ws.append(["A", "id1"])
    ws.append([None, None])
    ws.append([None, None])
    ws.append(["B", "id2"])
    ws.append([None, None])
    remove_empty_rows(ws)
    assert ws.max_row == 3
    assert ws.cell(2, 1).value == "A"
    assert ws.cell(3, 1).value == "B"


def test_adversarial_header_never_deleted():
    wb = Workbook()
    ws = wb.active
    ws.append([None, None])  # row 1: looks empty but is header
    ws.append(["A", "id1"])
    remove_empty_rows(ws)
    assert ws.max_row == 2  # row 1 kept


def test_adversarial_empty_string_not_treated_as_none():
    """Empty string '' is not None — row must NOT be deleted."""
    wb = Workbook()
    ws = wb.active
    ws.append(["标题", "ID"])
    ws.append(["", ""])          # row 2: both cells empty string, not None
    ws.append([None, None])      # row 3: real None — should be deleted
    remove_empty_rows(ws)
    assert ws.max_row == 2
    assert ws.cell(2, 1).value == ""
    assert ws.cell(2, 2).value == ""


def test_adversarial_col3_data_with_col1_col2_none_is_deleted():
    """Only col1+col2 determine emptiness; col3+ data is irrelevant per spec."""
    wb = Workbook()
    ws = wb.active
    ws.append(["标题", "ID", "数据"])
    ws.append([None, None, 999])  # col1/col2 None → delete regardless of col3
    ws.append(["A", "id1", 100])
    remove_empty_rows(ws)
    assert ws.max_row == 2
    assert ws.cell(2, 1).value == "A"


def test_adversarial_returns_none():
    """Function returns None (in-place mutation contract)."""
    wb = Workbook()
    ws = wb.active
    ws.append(["标题", "ID"])
    ws.append([None, None])
    result = remove_empty_rows(ws)
    assert result is None


def test_adversarial_header_only_sheet():
    """Header-only sheet: no data rows — must not raise and max_row stays 1."""
    wb = Workbook()
    ws = wb.active
    ws.append(["标题", "ID"])
    remove_empty_rows(ws)
    assert ws.max_row == 1
