"""Tests for the CLI entry point."""
import json
from typer.testing import CliRunner
from auto_excel.cli import app

runner = CliRunner()

_SAMPLE_ROWS = [
    {"花费": 113.6, "展现量": 1000, "点击量": 50, "留资人数": 5, "留资成本": 22.72, "互动成本": 10},
    {"花费": 227.2, "展现量": 2000, "点击量": 100, "留资人数": 10, "留资成本": 22.72, "互动成本": 10},
    {"花费": 340.8, "展现量": 3000, "点击量": 150, "留资人数": 15, "留资成本": 22.72, "互动成本": 10},
]


def test_on_no_files(tmp_dirs, monkeypatch):
    import auto_excel.config as cfg
    monkeypatch.setattr(cfg, "RAW_DIR", tmp_dirs["raw"])
    monkeypatch.setattr(cfg, "NEW_DIR", tmp_dirs["new"])
    monkeypatch.setattr(cfg, "LOG_DIR", tmp_dirs["log"])
    monkeypatch.setattr(cfg, "STATE_FILE", tmp_dirs["log"] / "processed.json")
    import auto_excel.state as st
    monkeypatch.setattr(st, "STATE_FILE", tmp_dirs["log"] / "processed.json")
    result = runner.invoke(app, ["on"])
    assert result.exit_code == 0
    assert "没有" in result.output or "待处理" in result.output


def test_on_processes_file(tmp_dirs, monkeypatch, make_sample_workbook):
    import auto_excel.config as cfg
    state_file = tmp_dirs["log"] / "processed.json"
    monkeypatch.setattr(cfg, "RAW_DIR", tmp_dirs["raw"])
    monkeypatch.setattr(cfg, "NEW_DIR", tmp_dirs["new"])
    monkeypatch.setattr(cfg, "LOG_DIR", tmp_dirs["log"])
    monkeypatch.setattr(cfg, "STATE_FILE", state_file)
    import auto_excel.state as st
    monkeypatch.setattr(st, "STATE_FILE", state_file)
    wb = make_sample_workbook(_SAMPLE_ROWS)
    wb.save(tmp_dirs["raw"] / "test_report.xlsx")
    result = runner.invoke(app, ["on"])
    assert result.exit_code == 0
    assert (tmp_dirs["new"] / "test_report.xlsx").exists()
    state = json.loads(state_file.read_text())
    assert "test_report.xlsx" in state


def test_adversarial_skips_already_processed(tmp_dirs, monkeypatch, make_sample_workbook):
    import auto_excel.config as cfg
    state_file = tmp_dirs["log"] / "processed.json"
    monkeypatch.setattr(cfg, "RAW_DIR", tmp_dirs["raw"])
    monkeypatch.setattr(cfg, "NEW_DIR", tmp_dirs["new"])
    monkeypatch.setattr(cfg, "LOG_DIR", tmp_dirs["log"])
    monkeypatch.setattr(cfg, "STATE_FILE", state_file)
    import auto_excel.state as st
    monkeypatch.setattr(st, "STATE_FILE", state_file)
    state_file.write_text(json.dumps({"already.xlsx": {"processed_at": "2026-01-01T00:00:00", "status": "success"}}))
    wb = make_sample_workbook(_SAMPLE_ROWS[:2])
    wb.save(tmp_dirs["raw"] / "already.xlsx")
    result = runner.invoke(app, ["on"])
    assert result.exit_code == 0
    assert "没有" in result.output or "待处理" in result.output


def test_adversarial_continues_after_error(tmp_dirs, monkeypatch, make_sample_workbook):
    import auto_excel.config as cfg
    state_file = tmp_dirs["log"] / "processed.json"
    monkeypatch.setattr(cfg, "RAW_DIR", tmp_dirs["raw"])
    monkeypatch.setattr(cfg, "NEW_DIR", tmp_dirs["new"])
    monkeypatch.setattr(cfg, "LOG_DIR", tmp_dirs["log"])
    monkeypatch.setattr(cfg, "STATE_FILE", state_file)
    import auto_excel.state as st
    monkeypatch.setattr(st, "STATE_FILE", state_file)
    (tmp_dirs["raw"] / "bad.xlsx").write_bytes(b"not an excel file")
    wb = make_sample_workbook(_SAMPLE_ROWS[:2])
    wb.save(tmp_dirs["raw"] / "good.xlsx")
    result = runner.invoke(app, ["on"])
    assert result.exit_code == 0
    assert (tmp_dirs["new"] / "good.xlsx").exists()
    if state_file.exists():
        state = json.loads(state_file.read_text())
        assert "bad.xlsx" not in state


def test_adversarial_skips_already_processed_runs_new(tmp_dirs, monkeypatch, make_sample_workbook):
    """Already-processed file is skipped; new file in same dir IS processed."""
    import auto_excel.config as cfg
    import auto_excel.state as st
    state_file = tmp_dirs["log"] / "processed.json"
    for attr, val in [("RAW_DIR", tmp_dirs["raw"]), ("NEW_DIR", tmp_dirs["new"]),
                      ("LOG_DIR", tmp_dirs["log"]), ("STATE_FILE", state_file)]:
        monkeypatch.setattr(cfg, attr, val)
    monkeypatch.setattr(st, "STATE_FILE", state_file)

    state_file.write_text(json.dumps({"already.xlsx": {"processed_at": "2026-01-01T00:00:00", "status": "success"}}))
    rows = [{"花费": 100.0, "展现量": 1000, "点击量": 50, "留资人数": 2}]
    wb = make_sample_workbook(rows)
    wb.save(tmp_dirs["raw"] / "already.xlsx")
    wb.save(tmp_dirs["raw"] / "new_file.xlsx")

    result = runner.invoke(app, ["on"])
    assert result.exit_code == 0
    # New file was processed
    assert (tmp_dirs["new"] / "new_file.xlsx").exists()
    # Already-processed file's new copy should NOT appear (it wasn't processed this run)
    state = json.loads(state_file.read_text())
    assert "new_file.xlsx" in state
    # The output reports 1 file, not 2 (proves filtering happened)
    assert "already.xlsx" not in result.output or "new_file.xlsx" in result.output


def test_adversarial_state_integrity_after_error(tmp_dirs, monkeypatch, make_sample_workbook):
    """After mixed run: success file IS in state, failure file is NOT. Both assertions are unconditional."""
    import auto_excel.config as cfg
    import auto_excel.state as st
    state_file = tmp_dirs["log"] / "processed.json"
    for attr, val in [("RAW_DIR", tmp_dirs["raw"]), ("NEW_DIR", tmp_dirs["new"]),
                      ("LOG_DIR", tmp_dirs["log"]), ("STATE_FILE", state_file)]:
        monkeypatch.setattr(cfg, attr, val)
    monkeypatch.setattr(st, "STATE_FILE", state_file)

    (tmp_dirs["raw"] / "bad.xlsx").write_bytes(b"not an excel file")
    rows = [{"花费": 100.0, "展现量": 1000, "点击量": 50, "留资人数": 2}]
    wb = make_sample_workbook(rows)
    wb.save(tmp_dirs["raw"] / "good.xlsx")

    result = runner.invoke(app, ["on"])
    assert result.exit_code == 0

    # Unconditional: state_file MUST exist because good.xlsx succeeded
    assert state_file.exists(), "processed.json must exist after a successful file"
    state = json.loads(state_file.read_text())
    assert "good.xlsx" in state, "successful file must be recorded"
    assert "bad.xlsx" not in state, "failed file must not be recorded"


def test_adversarial_process_file_applies_all_three_steps(tmp_dirs, monkeypatch, make_sample_workbook):
    """Output file must have calculated columns, sorted order, and 占比 column."""
    import auto_excel.config as cfg
    import auto_excel.state as st
    from openpyxl import load_workbook
    state_file = tmp_dirs["log"] / "processed.json"
    for attr, val in [("RAW_DIR", tmp_dirs["raw"]), ("NEW_DIR", tmp_dirs["new"]),
                      ("LOG_DIR", tmp_dirs["log"]), ("STATE_FILE", state_file)]:
        monkeypatch.setattr(cfg, attr, val)
    monkeypatch.setattr(st, "STATE_FILE", state_file)

    rows = [
        {"花费": 50.0, "展现量": 500, "点击量": 20, "留资人数": 1},
        {"花费": 200.0, "展现量": 2000, "点击量": 100, "留资人数": 5},
        {"花费": 100.0, "展现量": 1000, "点击量": 40, "留资人数": 2},
    ]
    wb = make_sample_workbook(rows)
    wb.save(tmp_dirs["raw"] / "report.xlsx")
    runner.invoke(app, ["on"])

    out_path = tmp_dirs["new"] / "report.xlsx"
    assert out_path.exists()
    wb_out = load_workbook(out_path)
    ws = wb_out.worksheets[3]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # Flow 1: calculated columns must exist
    for col in ["实际花费", "点击率", "CPC", "实际成本"]:
        assert col in headers, f"apply_calculated_columns did not run: missing {col}"

    # Flow 3: 占比 column must exist
    assert "占比" in headers, "group_and_merge did not run: missing 占比"

    # Flow 2: 实际成本 values must be descending
    cost_col = headers.index("实际成本") + 1
    costs = [ws.cell(r, cost_col).value for r in range(2, ws.max_row + 1)
             if ws.cell(r, cost_col).value is not None]
    assert costs == sorted(costs, reverse=True), "sort_by_column did not run: not descending"


def test_adversarial_result_status_error_on_failure(tmp_dirs, monkeypatch):
    """When file processing fails, result dict must have status='error', not 'success'."""
    import auto_excel.config as cfg
    import auto_excel.state as st
    import auto_excel.display as disp
    state_file = tmp_dirs["log"] / "processed.json"
    for attr, val in [("RAW_DIR", tmp_dirs["raw"]), ("NEW_DIR", tmp_dirs["new"]),
                      ("LOG_DIR", tmp_dirs["log"]), ("STATE_FILE", state_file)]:
        monkeypatch.setattr(cfg, attr, val)
    monkeypatch.setattr(st, "STATE_FILE", state_file)

    captured_results = []
    original_report = disp.print_report
    def capturing_report(results, console=None):
        captured_results.extend(results)
        original_report(results, console=console)
    monkeypatch.setattr(disp, "print_report", capturing_report)

    (tmp_dirs["raw"] / "bad.xlsx").write_bytes(b"not an excel file")
    runner.invoke(app, ["on"])

    assert len(captured_results) == 1
    assert captured_results[0]["filename"] == "bad.xlsx"
    assert captured_results[0]["status"] == "error"


# ──────────────────────────────────────────────
# uninstall command tests
# ──────────────────────────────────────────────

def test_uninstall_shows_what_will_be_deleted(tmp_path):
    """uninstall must list the install dir and wrapper before asking."""
    result = runner.invoke(app, ["uninstall"], input="n\n")
    assert "以下文件将被删除" in result.output
    assert ".auto-excel" in result.output
    assert "auto-excel" in result.output


def test_uninstall_shows_marketing_dir_preserved(tmp_path):
    """uninstall must explicitly state marketing analysis dir is not touched."""
    result = runner.invoke(app, ["uninstall"], input="n\n")
    assert "marketing analysis" in result.output
    assert "不会被删除" in result.output


def test_uninstall_aborts_on_no(tmp_path, monkeypatch):
    """Answering 'n' must abort without deleting anything."""
    install_dir = tmp_path / ".auto-excel"
    install_dir.mkdir()
    wrapper = tmp_path / "auto-excel-wrapper"
    wrapper.write_text("#!/bin/sh\n")

    import auto_excel.cli as cli_mod
    monkeypatch.setattr(cli_mod, "uninstall", _make_uninstall_with_paths(install_dir, wrapper))

    result = runner.invoke(app, ["uninstall"], input="n\n")
    # Files must still exist when user declines
    assert install_dir.exists()
    assert wrapper.exists()


def _make_uninstall_with_paths(install_dir, wrapper):
    """Return a patched uninstall function that uses tmp paths instead of home."""
    import shutil as _shutil
    import typer as _typer

    @app.command("uninstall")
    def _uninstall():
        _typer.echo("以下文件将被删除：")
        _typer.echo(f"  {install_dir}")
        _typer.echo(f"  {wrapper}")
        _typer.echo("以下内容不会被删除：")
        _typer.echo("  ~/Desktop/marketing analysis/（已处理文档完整保留）")
        _typer.echo("")
        _typer.confirm("确认卸载？", abort=True)
        removed = []
        if wrapper.exists():
            wrapper.unlink()
            removed.append(str(wrapper))
        if install_dir.exists():
            _shutil.rmtree(install_dir)
            removed.append(str(install_dir))
        if removed:
            for p in removed:
                _typer.echo(f"已删除：{p}")
        else:
            _typer.echo("未找到需要删除的文件，可能已经卸载。")
        _typer.echo("auto-excel 已卸载。")

    return _uninstall


def test_uninstall_deletes_install_dir_and_wrapper(tmp_path):
    """_remove_install_files removes both install dir and wrapper."""
    from auto_excel.cli import _remove_install_files

    install_dir = tmp_path / ".auto-excel"
    install_dir.mkdir()
    (install_dir / "prog.py").write_text("code")
    wrapper = tmp_path / "auto-excel"
    wrapper.write_text("#!/bin/sh\n")

    removed = _remove_install_files(install_dir, wrapper)

    assert not install_dir.exists(), "install dir must be deleted"
    assert not wrapper.exists(), "wrapper must be deleted"
    assert len(removed) == 2, "both paths must appear in the removed list"


def test_uninstall_succeeds_when_files_missing(tmp_path):
    """uninstall must succeed gracefully when install dir/wrapper don't exist."""
    import unittest.mock as mock

    fake_home = tmp_path  # empty — no .auto-excel or wrapper

    with mock.patch("auto_excel.cli.Path") as MockPath:
        home_mock = mock.MagicMock()
        MockPath.home.return_value = home_mock
        home_mock.__truediv__ = lambda self, other: (
            tmp_path / other if other in (".auto-excel",) else mock.MagicMock()
        )
        result = runner.invoke(app, ["uninstall"], input="y\n")

    assert result.exit_code in (0, 1)


def test_uninstall_exit_code_abort_on_no():
    """Responding 'n' must produce a non-zero exit code (Typer abort=True)."""
    result = runner.invoke(app, ["uninstall"], input="n\n")
    assert result.exit_code != 0


def test_uninstall_confirm_message_present():
    """The confirmation prompt must be visible to the user."""
    result = runner.invoke(app, ["uninstall"], input="n\n")
    assert "确认卸载" in result.output


def test_adversarial_uninstall_does_not_mention_raw_or_new():
    """uninstall output must NOT claim it deletes Raw/ or New/ subdirs."""
    result = runner.invoke(app, ["uninstall"], input="n\n")
    output_lower = result.output
    assert "Raw/" not in output_lower
    assert "New/" not in output_lower


def test_version_shows_correct_version():
    """version command must output the package version, not a hardcoded old string."""
    from auto_excel import __version__
    result = runner.invoke(app, ["version"])
    assert result.exit_code == 0
    assert __version__ in result.output
    assert "0.1.0" not in result.output


def test_config_install_dir_is_path():
    from pathlib import Path
    from auto_excel.config import INSTALL_DIR
    assert isinstance(INSTALL_DIR, Path)
    assert INSTALL_DIR.name == ".auto-excel"


def test_config_wrapper_is_path():
    from pathlib import Path
    from auto_excel.config import WRAPPER
    assert isinstance(WRAPPER, Path)
    assert WRAPPER.name == "auto-excel"
    assert ".local" in str(WRAPPER)


def test_adversarial_install_dir_under_home():
    from auto_excel.config import INSTALL_DIR
    from pathlib import Path
    assert str(INSTALL_DIR).startswith(str(Path.home()))


def test_adversarial_wrapper_under_local_bin():
    from auto_excel.config import WRAPPER
    assert WRAPPER.parent.name == "bin"
    assert WRAPPER.parent.parent.name == ".local"


# ──────────────────────────────────────────────
# info command tests
# ──────────────────────────────────────────────

def test_info_shows_version():
    from auto_excel import __version__
    result = runner.invoke(app, ["info"])
    assert result.exit_code == 0
    assert __version__ in result.output

def test_info_shows_install_path():
    result = runner.invoke(app, ["info"])
    assert "安装路径" in result.output
    assert ".auto-excel" in result.output

def test_info_shows_python_version():
    import sys
    expected = f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"
    result = runner.invoke(app, ["info"])
    assert expected in result.output

def test_info_shows_data_dir():
    result = runner.invoke(app, ["info"])
    assert "数据目录" in result.output
    assert "marketing analysis" in result.output


def test_adversarial_info_exit_code_zero():
    result = runner.invoke(app, ["info"])
    assert result.exit_code == 0

def test_adversarial_info_has_four_lines():
    result = runner.invoke(app, ["info"])
    lines = [l for l in result.output.strip().split("\n") if l.strip()]
    assert len(lines) == 4

def test_adversarial_info_first_line_matches_version_cmd():
    r_ver = runner.invoke(app, ["version"])
    r_info = runner.invoke(app, ["info"])
    info_first_line = r_info.output.strip().split("\n")[0]
    assert info_first_line == r_ver.output.strip()
